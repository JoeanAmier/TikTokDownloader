import csv
import logging
import os
import time

from openpyxl import Workbook
from openpyxl import load_workbook

from StringCleaner import Cleaner


class RunLogger:
    def __init__(self):
        self.log = None  # 记录器主体
        self._root = "./"  # 日志记录保存根路径
        self._folder = "Log"  # 日志记录保存文件夹名称
        self._name = "%Y-%m-%d %H.%M.%S"  # 日志文件名称

    @property
    def root(self):
        return self._root

    @root.setter
    def root(self, value):
        if os.path.exists(value) and os.path.isdir(value):
            self._root = value
        else:
            print("日志保存路径错误！将使用当前路径作为日志保存路径！")
            self._root = "./"

    @property
    def name(self):
        return self._name

    @name.setter
    def name(self, value):
        if value:
            try:
                _ = time.strftime(value, time.localtime())
                self._name = value
            except ValueError:
                print("日志名称格式错误，将使用默认时间格式（年-月-日 时.分.秒）")
                self._name = "%Y-%m-%d %H.%M.%S"
        else:
            print("日志名称格式错误，将使用默认时间格式（年-月-日 时.分.秒）")
            self._name = "%Y-%m-%d %H.%M.%S"

    @property
    def folder(self):
        return self._folder

    @folder.setter
    def folder(self, value):
        """未生效"""
        pass

    def run(
            self,
            format_="%(asctime)s[%(levelname)s]%(filename)s-%(lineno)d: %(message)s"):
        if not os.path.exists(dir_ := os.path.join(self.root, self.folder)):
            os.mkdir(dir_)
        self.log = logging
        self.log.basicConfig(
            filename=os.path.join(
                dir_,
                f"{time.strftime(self.name, time.localtime())}.log"),
            level=logging.INFO,
            datefmt='%Y-%m-%d %H:%M:%S',
            format=format_,
            encoding="UTF-8")

    def info(self, text: str, output=True):
        if output:
            print(text)
        self.log.info(text)

    def warning(self, text: str, output=True):
        if output:
            print(text)
        self.log.warning(text)

    def error(self, text: str, output=True):
        if output:
            print(text)
        self.log.error(text)


class NoneLogger:
    def __init__(self, *args, **kwargs):
        pass

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        pass

    def save(self, *args, **kwargs):
        pass


class CSVLogger:
    """CSV格式记录"""

    def __init__(self, root: str, name="Download"):
        self.file = None  # 文件对象
        self.writer = None  # CSV对象
        self.root = root  # 文件路径
        self.name = name  # 文件名称

    def __enter__(self):
        if not os.path.exists(self.root):
            os.mkdir(self.root)
        self.root = os.path.join(self.root, f"{self.name}.csv")
        self.file = open(self.root,
                         "a",
                         encoding="UTF-8",
                         newline="")
        self.writer = csv.writer(self.file)
        self.title()
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.file.close()

    def title(self):
        if os.path.getsize(self.root) == 0:
            # 如果文件没有任何数据，则写入标题行
            self.save(RecordManager.title)

    def save(self, data):
        self.writer.writerow(data)


class XLSXLogger:
    """XLSX格式"""

    def __init__(self, root: str, name="Download"):
        self.book = None  # XLSX数据簿
        self.sheet = None  # XLSX数据表
        self.root = root  # 文件路径
        self.name = name  # 文件名称

    def __enter__(self):
        if not os.path.exists(self.root):
            os.mkdir(self.root)
        self.root = os.path.join(self.root, f"{self.name}.xlsx")
        if os.path.exists(self.root):
            self.book = load_workbook(self.root)
        else:
            self.book = Workbook()
        self.sheet = self.book.active
        self.title()
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.book.save(self.root)
        self.book.close()

    def title(self):
        if not self.sheet["A1"].value:
            # 如果文件没有任何数据，则写入标题行
            for col, value in enumerate(RecordManager.title, start=1):
                self.sheet.cell(row=1, column=col, value=value)

    def save(self, data):
        self.sheet.append(data)


class RecordManager:
    """检查数据记录路径"""
    title = ("作品类型", "作品ID", "作品描述", "发布时间", "账号昵称", "Video_ID")

    @staticmethod
    def run(root="./", folder="Data"):
        if not os.path.exists(root):
            return False
        return os.path.join(
            root, r) if (
            r := Cleaner().filter(folder)) else False
