from pathlib import Path
from typing import TYPE_CHECKING

from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import IllegalCharacterError

from ..translation import _
from .text import BaseTextLogger

if TYPE_CHECKING:
    from ..tools import ColorfulConsole

__all__ = ["XLSXLogger"]


class XLSXLogger(BaseTextLogger):
    """XLSX 格式保存数据"""

    __type = "xlsx"

    def __init__(
        self,
        root: Path,
        title_line: tuple,
        field_keys: tuple,
        console: "ColorfulConsole",
        old=None,
        name="Download",
        *args,
        **kwargs,
    ):
        super().__init__(*args, **kwargs)
        self.console = console
        self.book = None  # XLSX数据簿
        self.sheet = None  # XLSX数据表
        self.name = self._rename(root, self.__type, old, name)  # 文件名称
        self.path = root.joinpath(f"{self.name}.{self.__type}")
        self.title_line = title_line  # 标题行
        self.field_keys = field_keys

    async def __aenter__(self):
        self.book = load_workbook(self.path) if self.path.exists() else Workbook()
        self.sheet = self.book.active
        self.title()
        return self

    async def __aexit__(self, exc_type, exc_val, exc_tb):
        self.book.save(self.path)
        self.book.close()

    def title(self):
        if not self.sheet["A1"].value:
            # 如果文件没有任何数据，则写入标题行
            for col, value in enumerate(self.title_line, start=1):
                self.sheet.cell(row=1, column=col, value=value)

    async def _save(self, data, *args, **kwargs):
        try:
            self.sheet.append(data)
        except IllegalCharacterError as e:
            self.console.warning(
                _("数据包含非法字符，保存数据失败：{error}").format(error=e)
            )
